# apps/solicitudes/services/generar_reportes.py

import io
from datetime import timedelta

from django.utils import timezone
from django.db.models import Sum, Count, Q


# ------------------------------------------------
# HELPERS — OBTENER CONSUMIDORES SEGÚN FILTRO
# ------------------------------------------------

def _get_consumidores(filtro: str, dias: int = 30):
    """
    Retorna queryset de ConsumidorPerfil según el filtro:
      - BLOQUEADOS       → alerta_repetitividad=BLOQUEADO
      - EN_REVISION      → alerta_repetitividad=EN_REVISION
      - SUPERARON_LIMITE → superaron el límite en X días
      - TODOS            → todos los consumidores
    """
    from consumidores.models import ConsumidorPerfil
    from configuracion.models import ConfiguracionSistema
    from solicitudes.models import Solicitud

    config = ConfiguracionSistema.obtener()
    qs = ConsumidorPerfil.objects.select_related(
        "user", "departamento", "provincia", "municipio"
    ).prefetch_related("documentos")

    if filtro == "BLOQUEADOS":
        return qs.filter(
            alerta_repetitividad=ConsumidorPerfil.EstadoAlerta.BLOQUEADO
        )

    if filtro == "EN_REVISION":
        return qs.filter(
            alerta_repetitividad=ConsumidorPerfil.EstadoAlerta.EN_REVISION
        )

    if filtro == "SUPERARON_LIMITE":
        desde = timezone.now() - timedelta(days=dias)
        ids = Solicitud.objects.filter(
            estado__in=["APROBADA", "DESPACHADA"],
            fecha_creacion__gte=desde
        ).values("consumidor").annotate(
            total=Count("id")
        ).filter(
            total__gte=config.limite_solicitudes_por_periodo
        ).values_list("consumidor", flat=True)
        return qs.filter(id__in=ids)

    # TODOS
    return qs


def _get_datos_consumidor(perfil, dias: int = 30):
    """
    Retorna los datos completos de un consumidor
    para incluir en el reporte.
    """
    from solicitudes.models import Solicitud

    desde = timezone.now() - timedelta(days=dias)

    solicitudes = Solicitud.objects.filter(
        consumidor=perfil,
        estado__in=["APROBADA", "DESPACHADA"],
        fecha_creacion__gte=desde
    ).select_related("estacion_servicio").order_by("fecha_creacion")

    totales = solicitudes.aggregate(
        total_litros_solicitados=Sum("litros_solicitados"),
        total_litros_despachados=Sum("litros_despachados"),
    )

    doc = perfil.documentos.first()

    return {
        "nombre_completo": perfil.user.nombre_completo(),
        "email":           perfil.user.email,
        "ci":              f"{doc.numero_documento} {doc.complemento_documento}".strip() if doc else "—",
        "tipo_doc":        doc.get_tipo_documento_display() if doc else "—",
        "departamento":    perfil.departamento.nombre if perfil.departamento else "—",
        "municipio":       perfil.municipio.nombre    if perfil.municipio    else "—",
        "direccion":       perfil.direccion or "—",
        "alerta":          perfil.get_alerta_repetitividad_display(),
        "motivo":          perfil.motivo_bloqueo or "—",
        "total_solicitudes":        solicitudes.count(),
        "total_litros_solicitados": totales["total_litros_solicitados"] or 0,
        "total_litros_despachados": totales["total_litros_despachados"] or 0,
        "solicitudes": [
            {
                "id_publico":    str(s.id_publico)[:8].upper(),
                "fecha":         s.fecha_creacion.strftime("%d/%m/%Y"),
                "fecha_despacho": s.fecha_despacho.strftime("%d/%m/%Y") if s.fecha_despacho else "—",
                "litros_sol":    s.litros_solicitados,
                "litros_des":    s.litros_despachados or 0,
                "combustible":   s.get_tipo_combustible_display(),
                "estacion":      s.estacion_servicio.nombre if s.estacion_servicio else "—",
            }
            for s in solicitudes
        ]
    }


# ------------------------------------------------
# GENERAR REPORTE EXCEL
# ------------------------------------------------

def generar_reporte_excel(filtro: str, dias: int = 30) -> bytes:
    """
    Genera un archivo Excel con el reporte de consumidores.
    Retorna bytes del archivo.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Consumidores"

    # Estilos
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="1a3a5c")
    subheader_fill = PatternFill("solid", fgColor="2d6a9f")
    alerta_fill    = PatternFill("solid", fgColor="FFC7CE")
    center         = Alignment(horizontal="center", vertical="center")
    thin_border    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    # Título
    ws.merge_cells("A1:K1")
    ws["A1"] = f"REPORTE ANH — {filtro.replace('_', ' ')} — {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = header_fill
    ws["A1"].alignment = center

    # Cabeceras
    headers = [
        "Nombre completo", "Tipo Doc.", "N° Documento",
        "Email", "Departamento", "Municipio", "Dirección",
        "Total solicitudes", "Litros solicitados",
        "Litros despachados", "Estado alerta"
    ]

    for col, header in enumerate(headers, 1):
        cell            = ws.cell(row=2, column=col, value=header)
        cell.font       = header_font
        cell.fill       = subheader_fill
        cell.alignment  = center
        cell.border     = thin_border

    # Datos
    consumidores = _get_consumidores(filtro, dias)
    row = 3

    for perfil in consumidores:
        datos = _get_datos_consumidor(perfil, dias)

        valores = [
            datos["nombre_completo"],
            datos["tipo_doc"],
            datos["ci"],
            datos["email"],
            datos["departamento"],
            datos["municipio"],
            datos["direccion"],
            datos["total_solicitudes"],
            datos["total_litros_solicitados"],
            datos["total_litros_despachados"],
            datos["alerta"],
        ]

        for col, valor in enumerate(valores, 1):
            cell           = ws.cell(row=row, column=col, value=valor)
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center")

            # Resaltar consumidores bloqueados
            if datos["alerta"] == "Bloqueado":
                cell.fill = alerta_fill

        row += 1

    # Ajustar anchos de columna
    anchos = [25, 12, 18, 28, 15, 15, 25, 10, 12, 12, 20]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # Guardar en memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ------------------------------------------------
# GENERAR REPORTE PDF
# ------------------------------------------------

def generar_reporte_pdf(filtro: str, dias: int = 30) -> bytes:
    """
    Genera un PDF con el reporte detallado de consumidores.
    Incluye tabla resumen y detalle de solicitudes por consumidor.
    Retorna bytes del archivo.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm,
    )

    styles  = getSampleStyleSheet()
    azul    = colors.HexColor("#1a3a5c")
    azul2   = colors.HexColor("#2d6a9f")
    rojo    = colors.HexColor("#c0392b")
    gris    = colors.HexColor("#f2f2f2")
    blanco  = colors.white

    titulo_style = ParagraphStyle(
        "titulo", fontSize=16, textColor=blanco,
        alignment=TA_CENTER, fontName="Helvetica-Bold"
    )
    subtitulo_style = ParagraphStyle(
        "subtitulo", fontSize=11, textColor=azul,
        alignment=TA_CENTER, fontName="Helvetica-Bold",
        spaceAfter=6
    )
    normal_style = ParagraphStyle(
        "normal", fontSize=9, fontName="Helvetica",
        textColor=colors.black
    )
    bold_style = ParagraphStyle(
        "bold", fontSize=9, fontName="Helvetica-Bold",
        textColor=azul
    )

    elementos = []

    # ------------------------------------------------
    # ENCABEZADO
    # ------------------------------------------------

    encabezado = Table(
        [[Paragraph(
            f"AGENCIA NACIONAL DE HIDROCARBUROS — BOLIVIA<br/>"
            f"Reporte de Consumidores: {filtro.replace('_', ' ')}<br/>"
            f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')} | Período: {dias} días",
            titulo_style
        )]],
        colWidths=["100%"]
    )
    encabezado.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), azul),
        ("TOPPADDING",    (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
    ]))
    elementos.append(encabezado)
    elementos.append(Spacer(1, 0.4*cm))

    # ------------------------------------------------
    # TABLA RESUMEN
    # ------------------------------------------------

    consumidores = _get_consumidores(filtro, dias)

    elementos.append(Paragraph("RESUMEN DE CONSUMIDORES", subtitulo_style))

    cabeceras = [
        "Nombre completo", "Documento", "Email",
        "Municipio", "Solicitudes", "Lit. Solicitados",
        "Lit. Despachados", "Alerta"
    ]

    filas = [cabeceras]

    for perfil in consumidores:
        datos = _get_datos_consumidor(perfil, dias)
        filas.append([
            datos["nombre_completo"],
            f"{datos['tipo_doc']}\n{datos['ci']}",
            datos["email"],
            f"{datos['municipio']}\n{datos['departamento']}",
            str(datos["total_solicitudes"]),
            f"{datos['total_litros_solicitados']} L",
            f"{datos['total_litros_despachados']} L",
            datos["alerta"],
        ])

    tabla = Table(filas, repeatRows=1)
    tabla.setStyle(TableStyle([
        # Encabezado
        ("BACKGROUND",   (0, 0), (-1, 0),  azul2),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  blanco),
        ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0),  8),
        ("ALIGN",        (0, 0), (-1, 0),  "CENTER"),
        ("TOPPADDING",   (0, 0), (-1, 0),  6),
        ("BOTTOMPADDING",(0, 0), (-1, 0),  6),
        # Filas
        ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",     (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [blanco, gris]),
        ("ALIGN",        (4, 1), (-1, -1), "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",   (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 1), (-1, -1), 4),
        # Bordes
        ("GRID",         (0, 0), (-1, -1), 0.5, colors.grey),
        ("BOX",          (0, 0), (-1, -1), 1,   azul),
    ]))

    # Colorear filas de bloqueados
    for i, perfil in enumerate(consumidores, 1):
        if perfil.alerta_repetitividad == "BLOQUEADO":
            tabla.setStyle(TableStyle([
                ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FFCCCC")),
                ("TEXTCOLOR",  (7, i), (7,  i), rojo),
                ("FONTNAME",   (7, i), (7,  i), "Helvetica-Bold"),
            ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 0.5*cm))

    # ------------------------------------------------
    # DETALLE POR CONSUMIDOR
    # ------------------------------------------------

    elementos.append(HRFlowable(width="100%", thickness=1, color=azul))
    elementos.append(Spacer(1, 0.3*cm))
    elementos.append(Paragraph("DETALLE DE SOLICITUDES POR CONSUMIDOR", subtitulo_style))

    for perfil in consumidores:
        datos = _get_datos_consumidor(perfil, dias)

        if not datos["solicitudes"]:
            continue

        # Encabezado del consumidor
        info = Table([[
            Paragraph(f"<b>{datos['nombre_completo']}</b>", bold_style),
            Paragraph(f"Doc: {datos['tipo_doc']} {datos['ci']}", normal_style),
            Paragraph(f"Email: {datos['email']}", normal_style),
            Paragraph(f"Alerta: {datos['alerta']}", bold_style),
        ]], colWidths=["30%", "20%", "30%", "20%"])
        info.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), gris),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("BOX",           (0, 0), (-1, -1), 0.5, azul),
        ]))
        elementos.append(info)

        # Tabla de solicitudes del consumidor
        cab_sol = [
            "N° Solicitud", "Fecha solicitud", "Fecha despacho",
            "Combustible", "Lit. Sol.", "Lit. Des.", "Estación"
        ]
        filas_sol = [cab_sol]
        for s in datos["solicitudes"]:
            filas_sol.append([
                s["id_publico"],
                s["fecha"],
                s["fecha_despacho"],
                s["combustible"],
                f"{s['litros_sol']} L",
                f"{s['litros_des']} L",
                s["estacion"],
            ])

        tabla_sol = Table(filas_sol, repeatRows=1)
        tabla_sol.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  azul2),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  blanco),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 7.5),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [blanco, gris]),
            ("ALIGN",         (4, 0), (-1, -1), "CENTER"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("GRID",          (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elementos.append(tabla_sol)
        elementos.append(Spacer(1, 0.4*cm))

    doc.build(elementos)
    buffer.seek(0)
    return buffer.getvalue()